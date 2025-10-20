import pandas as pd
from io import BytesIO
from datetime import datetime, date
from openpyxl import Workbook

def str_para_data(data):
    """Converte vários tipos para datetime.date ou retorna None."""
    # trata NaN/None
    if pd.isna(data):
        return None

    # datetime / date / pandas.Timestamp
    if isinstance(data, datetime):
        return data.date()
    if isinstance(data, date) and not isinstance(data, pd.Timestamp):
        return data
    if isinstance(data, pd.Timestamp):
        return data.date()

    # tenta converter string / numeros / numpy datetime com pandas
    try:
        ts = pd.to_datetime(data, dayfirst=True, errors='coerce')
        if pd.isna(ts):
            return None
        return ts.date()
    except Exception:
        return None

def tempo_atraso(data):
    data_conv = str_para_data(data)
    if data_conv is None:
        return None
    hoje = datetime.today().date()
    return (hoje - data_conv).days

def classificacao(tempo):
    if tempo is None:
        return "Sem Data"
    if tempo <= 0:
        return "No Prazo"
    if tempo < 30:
        return "A"
    elif tempo < 45:
        return "B"
    elif tempo < 60:
        return "C"
    elif tempo < 90:
        return "D"
    else:
        return "E"

def tipo(UM):
    match UM:
        case "SV" | "UA":
            return "Serviço"
        case _:
            return "Material"

def localidade(grupo):
    local = {
        120: 'SP', 128: 'MG', 129: 'MG', 130: 'MG',
        132: 'MG', 133: 'SP', 134: 'SP', 135: 'SP', 136: 'MG',
        137: 'MG', 138: 'MG', 139: 'MG', 141: 'MG',
        142: 'MG', 143: 'MG', 144: 'MG', 146: 'SP'
    }
    return local.get(grupo, 'N/A')

def g_comprador(grupo):
    compradores = {
        120: 'Alexsandro', 128: 'João', 129: 'Diane', 130: 'Dinora',
        132: 'Palloma',133: 'Barbara', 134: 'Luci', 135: 'Roberta',
        136: 'Diego', 137: 'Eliene', 138: 'Felipe',
        139: 'Rafael', 141: 'Gabriel',142: 'Lucimara', 143: 'Tiago',
        144: 'Andre', 146: 'Fabiana'
    }
    return compradores.get(grupo, 'N/A')

def conversor(arquivo_csv):

    # Cria um novo Excel na memória
    arquivo = Workbook()
    aba_atual = arquivo.active
    aba_atual.title = "Base Atrasados"

    # Cabeçalhos
    colunas = [
        "Data do Documento","Documento de compras","Item","Texto Breve",
        "Data de Remessa","Fornecedor","Requisição de compras","Item ReqC","Comprador",
        "Tempo de atraso", "Classificação","Chave","Tipo","Status","Localidade"
    ]

    aba_atual.append(colunas)

    base = pd.read_csv(arquivo_csv, sep=';', encoding='latin1')

    data_dict = {}
    pedido_dict = {}
    item_dict = {}
    material_dict = {}
    remessa_dict = {}
    fornecedor_dict = {}
    requisicao_dict = {}
    itemrc_dict = {}
    grupo_dict = {}
    ua_dict = {}

    for indice, linha in base.iterrows():
        data = linha.get('Data do documento', None)
        pedido = linha.get('Documento de compras', '')
        item = linha.get('Item', '')
        material = linha.get('Texto breve', '')
        remessa = linha.get('Data de remessa', None)
        fornecedor = linha.get('Fornecedor/centro fornecedor', '')
        requisicao = linha.get('Requisição de compra', '')
        item_rc = linha.get('Item RC', '')
        grupo = linha.get('Grupo de compradores', None)
        ua = linha.get('UM pedido', '')

        codigo = str(pedido) + str(item)  # chave consistente como string
        aba_atual.cell(row=indice + 2, column=12).value = codigo

        # armazena apenas se nao existir
        data_dict.setdefault(codigo, data)
        pedido_dict.setdefault(codigo, pedido)
        item_dict.setdefault(codigo, item)
        material_dict.setdefault(codigo, material)
        remessa_dict.setdefault(codigo, remessa)
        fornecedor_dict.setdefault(codigo, fornecedor)
        requisicao_dict.setdefault(codigo, requisicao)
        itemrc_dict.setdefault(codigo, item_rc)   # corrigido: usar item_rc
        grupo_dict.setdefault(codigo, grupo)
        ua_dict.setdefault(codigo, ua)

    # preenche a aba lendo as linhas existentes
    # iter_rows retorna tuplas; verifica codigo seguro
    for indice, linha in enumerate(aba_atual.iter_rows(min_row=2, values_only=True), start=2):
        codigo = linha[11]  # coluna 12 (index 11)
        if not codigo:
            # se nao tem codigo, pula
            continue
        codigo = str(codigo)

        if codigo in data_dict:
            aba_atual.cell(row=indice, column=1).value = data_dict[codigo]
            aba_atual.cell(row=indice, column=2).value = pedido_dict[codigo]
            aba_atual.cell(row=indice, column=3).value = item_dict[codigo]
            aba_atual.cell(row=indice, column=4).value = material_dict[codigo]
            aba_atual.cell(row=indice, column=5).value = remessa_dict[codigo]
            # fornecedor: limpo e, se quiser só 6 primeiros digitos use [:6]
            fornecedor_val = str(fornecedor_dict.get(codigo, '')).strip()
            aba_atual.cell(row=indice, column=6).value = fornecedor_val[11:] if len(fornecedor_val) > 11 else fornecedor_val
            aba_atual.cell(row=indice, column=7).value = requisicao_dict[codigo]
            aba_atual.cell(row=indice, column=8).value = itemrc_dict[codigo]
            aba_atual.cell(row=indice, column=9).value = g_comprador(grupo_dict.get(codigo))
            aba_atual.cell(row=indice, column=13).value = tipo(ua_dict.get(codigo))

        # calcula tempo e classificacao de forma segura
        tempo = remessa_dict.get(codigo, None)
        tempo = tempo_atraso(tempo) if tempo is not None else None
        aba_atual.cell(row=indice, column=10).value = tempo
        classif = classificacao(tempo)
        aba_atual.cell(row=indice, column=11).value = classif
        if classif == "No Prazo":
            aba_atual.cell(row=indice, column=14).value = "No prazo"
        elif classif == "Sem Data":
            aba_atual.cell(row=indice, column=14).value = "Sem Data"
        else:
            aba_atual.cell(row=indice, column=14).value = "Em Atraso"

        aba_atual.cell(row=indice, column=15).value = localidade(grupo_dict.get(codigo))

  # Salva na memória
    buffer = BytesIO()
    arquivo.save(buffer)
    buffer.seek(0)
    return buffer