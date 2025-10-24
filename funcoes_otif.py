import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from datetime import datetime,date

def str_para_data(data_str):
    """Converte string de data para um objeto datetime.date."""
    if isinstance(data_str, date):  
        return data_str  # Se já for um objeto datetime.date, retorna diretamente

    if not isinstance(data_str, str) or not data_str.strip():  
        return None  # Retorna None se for vazio, None ou não for string válida

    formatos = ["%Y-%m-%d", "%d/%m/%Y"]  # Suporte para múltiplos formatos

    for formato in formatos:
        try:
            return datetime.strptime(data_str, formato).date()
        except ValueError:
            continue  # Tenta o próximo formato se falhar

    return None  # Retorna None caso nenhum formato seja compatível
        
def converter_num(numero):
    """Converter números"""
    numero = numero.replace('.', '')  # Remove os pontos (separadores de milhar)
    numero = numero.replace(',', '.')  # Substitui vírgula por ponto (se necessário)
    numero = float(numero)
    return numero

def gerar_codigo(pedido,item):
        """Gerar código do item do pedido"""
        pedido = str(pedido)
        if pedido[:2] == '45':
            final_pedido = pedido[4:]
        elif pedido[:2] == '46':
            final_pedido = pedido[5:]
        elif pedido[:2] == '43':
            final_pedido = pedido[6:]
        codigo = f'{final_pedido}{item}'

        return codigo

def localidade(grupo):
    """Seperar grupo de compras de acordo com a localidade"""

    local= {
        120: 'SP', 128: 'MG', 129: 'MG', 130: 'MG',
        132: 'MG',133: 'SP', 134: 'SP', 135: 'SP',136: 'MG',
        137: 'MG', 138: 'MG', 139: 'MG', 141: 'MG',
        142: 'MG', 143: 'MG', 144: 'MG', 146: 'SP'
    }
    return local.get(grupo, 'N/A')

def g_comprador(grupo):
    """Classificar grupo de compras de acordo com o comprador"""

    compradores = {
        120: 'Alexsandro', 128: 'Leonardo', 129: 'Diane', 130: 'Dinora',
        132: 'Palloma',133: 'Barbara', 134: 'Luci', 135: 'Roberta',
        136: 'Diego', 137: 'Eliene', 138: 'Felipe',
        139: 'Rafael', 141: 'Gabriel',142: 'Lucimara', 143: 'Tiago',
        144: 'Andre', 146: 'Fabiana'
    }
    return compradores.get(grupo, 'N/A')

def inicio(me80fn, yb,me2n):

    df_me80fn = pd.read_csv(me80fn,sep=';',encoding='latin1',low_memory=False)
    df_yb =  pd.read_csv(yb,sep=';',encoding='latin1',low_memory=False)
    df_me2n = pd.read_csv(me2n,sep=';',encoding='latin1')
    
        # Cria um novo Excel na memória
    arquivo = Workbook()
    aba_atual = arquivo.active
    aba_atual.title = "Base"

      # Cabeçalhos
    colunas = [
        "Nº Pedido","Item","Final","Fornecedor",
        "comprador","Data Criação","Data de Remessa SAP","Frete","Faturamento",
        "Entrada Fisíca", "Entrada Fiscal","Qtde Solicitada","Qtde Entregue","Localidade"
    ]

    aba_atual.append(colunas)

    # 'Documento de compras', 'Item'
    for indice, linha in df_me80fn.iterrows():
        pedido = linha['Documento de compras']
        item = linha['Item']
        codigo = gerar_codigo(pedido,item)
        aba_atual.cell(row=indice + 2,column=1).value = pedido
        aba_atual.cell(row=indice + 2,column=2).value = item
        aba_atual.cell(row=indice + 2,column=3).value = int(codigo)

    dictQ, dictE,fornecedor_dict,comprador_dict,remessa_dict,data_dict,frete_dict,faturamento_dict,qtd_solicitadadict,qtd_entreguedict,extorno_dict, localidade_dict = {},{},{},{},{},{},{},{},{},{},{},{}

    for indice, linha in df_me80fn.iterrows():
        codigo = linha['Chave']
        ctg_pedido = linha['Ctg.de histórico de pedido']
        cod_movimento_Q = f'{codigo}Q'
        cod_movimento_E = f'{codigo}E'
        movimento = linha['Tipo de movimento']
        cod_debcred = linha['Cód.débito/crédito']
        data = linha['Data de entrada']
        faturamento = linha['Data do documento']
        qtd_entregue = linha['Quantidade']
        qtd_entregue = converter_num(qtd_entregue)


        if cod_movimento_Q not in dictQ:  # Verifique se a chave já existe
            dictQ[cod_movimento_Q] = data  # Atribua o valor de data à chave
        if cod_movimento_E not in dictE and movimento == 101:  # Verifique se a chave já existe
            dictE[cod_movimento_E] = data  # Atribua o valor de data à chave
        if codigo not in faturamento_dict:
            faturamento_dict[codigo] = faturamento
    

        if codigo in qtd_entreguedict and qtd_entregue > 0 and movimento != 102 and ctg_pedido =='Q':
                qtd_entreguedict[codigo] += qtd_entregue
        else:
            if codigo not in qtd_entreguedict and movimento != 102 and ctg_pedido =='Q':
                qtd_entreguedict[codigo] = qtd_entregue
    
        if (movimento == 102 and cod_debcred != 'H') or (cod_debcred =='H' and ctg_pedido =='Q'):
            if codigo in extorno_dict:
                extorno_dict[codigo] += qtd_entregue
            else:
                extorno_dict[codigo] = qtd_entregue

    for codigo in qtd_entreguedict:
        if codigo in extorno_dict:
            qtd_entreguedict[codigo] += extorno_dict[codigo]

    for indice, linha in df_me2n.iterrows():
        codigo = linha['Chave']
        remessa = linha['Dat.rem.estatística']
        qtd_solicitada = linha['Qtd.pedido']
        qtd_solicitada = converter_num(qtd_solicitada)

        if codigo not in qtd_solicitadadict:
            qtd_solicitadadict[codigo] = qtd_solicitada
        if codigo not in remessa_dict:
            remessa_dict[codigo] = remessa


    for indice, linha in df_yb.iterrows():
        codigo = linha['Chave']
        fornecedor = linha['Nº conta do fornecedor']
        comprador = linha['GCm']
        data_doc = linha['Data doc.']
        frete = linha['IncTm']


        if codigo not in fornecedor_dict:
            fornecedor_dict[codigo] = fornecedor
        if codigo not in comprador_dict:
            comprador_dict[codigo] = comprador
        if codigo not in remessa_dict:
            remessa_dict[codigo] = remessa
        if codigo not in data_dict:
            data_dict[codigo] = data_doc
        if codigo not in frete_dict:
            frete_dict[codigo] = frete

    for indice,linha in enumerate(aba_atual.iter_rows(values_only=True)):
        codigo = linha[2]
        cod_movimento_Q = f'{codigo}Q'
        cod_movimento_E = f'{codigo}E'

        if cod_movimento_Q in dictQ:
            dt_fiscal = dictQ[cod_movimento_Q]
            aba_atual.cell(row=indice + 1,column=11).value = dictQ[cod_movimento_Q]
        if cod_movimento_E in dictE:
            dt_fisica = dictE[cod_movimento_E]
            aba_atual.cell(row=indice + 1,column=10).value = dictE[cod_movimento_E]
        else:
            dt_fisica = ''
        if codigo in fornecedor_dict:
            aba_atual.cell(row=indice + 1,column=4).value = fornecedor_dict[codigo]
        if codigo in comprador_dict:
            aba_atual.cell(row=indice + 1,column=5).value = g_comprador(comprador_dict[codigo])
        if codigo in data_dict:
            aba_atual.cell(row=indice + 1,column=6).value = data_dict[codigo]
        if codigo in remessa_dict:
            dt_remessa = remessa_dict[codigo]

            aba_atual.cell(row=indice + 1,column=7).value = remessa_dict[codigo]
        else:
            dt_remessa = ''
        if codigo in frete_dict:
            frete = frete_dict[codigo]
            aba_atual.cell(row=indice + 1,column=8).value = frete_dict[codigo]
        if codigo in faturamento_dict:
            dt_faturamento = faturamento_dict[codigo]
            aba_atual.cell(row=indice + 1,column=9).value = faturamento_dict[codigo]
        if codigo in qtd_solicitadadict:
            aba_atual.cell(row=indice + 1,column=12).value = qtd_solicitadadict[codigo]
        if codigo in qtd_entreguedict:
            aba_atual.cell(row=indice + 1,column=13).value = qtd_entreguedict[codigo]
        if codigo in comprador_dict:
            aba_atual.cell(row=indice + 1,column=14).value = localidade(comprador_dict[codigo])
        # if codigo in localidade_dict:
        #     aba_atual.cell(row=indice + 1,column=16).value = localidade_dict[codigo]

        if indice > 0 and dt_remessa != '':
            if dt_fisica != '':
                dt_fisica = str_para_data(dt_fisica)
            dt_fiscal= str_para_data(dt_fiscal)
            dt_faturamento = str_para_data(dt_faturamento)
            dt_remessa = str_para_data(dt_remessa)
    buffer = BytesIO()
    arquivo.save(buffer)
    buffer.seek(0)
    return buffer